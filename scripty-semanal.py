import openpyxl as xl
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.styles.numbers import FORMAT_DATE_DDMMYY, FORMAT_NUMBER_COMMA_SEPARATED2
from openpyxl.styles.alignment import Alignment
from openpyxl.utils import get_column_letter
import pandas as pd

#--------------------------RELATÓRIO SEMANAL----------------------------------------------
#-------------------------------SUNHUB--------------------------------

# Dados SUNHUB
wbSun = xl.load_workbook("ArquivosSunhub\Relatório Diretoria - SUNHUB.xlsx")
wsSun = wbSun['Ploomes']

# Últimas Linhas e Colunas
maxR = wsSun.max_row
maxC = wsSun.max_column

dados = []
# Percorrendo os dados do SunHub
for row in wsSun.iter_rows(min_row=1, max_row=maxR, max_col=maxC):
    linha = []
    for cell in row:
        linha.append(cell.value)
    dados.append(linha)


#-------------------------------SEMANAL--------------------------------


# Criando a Workbook de Relatório
wbSem = xl.Workbook()
wsSem = wbSem.active
wsSem.title = "Semanal"

# Copiando os dados da Lista para a Workbook
for l, linha in enumerate(dados, start=1):  # Começa na linha 1
    for c, valor in enumerate(linha, start=1):  # Começa na coluna 1
        wsSem.cell(row=l, column=c, value=valor)


# LINHA 'TOTAL'
linha_total = wsSem.max_row
wsSem.cell(row=linha_total, column=1).value = "Total:"  # TOTAL
#wsSem.cell(row=linha_total, column=2).value = f'=CONT.VALORES(B2:B{linha_total - 1})'  # SOMA DE VENDAS
#wsSem.cell(row=linha_total, column=3).value = f'=SUM(C2:C{linha_total - 1})'  # SOMA DOS VALORES
wsSem.cell(row=linha_total, column=2).alignment = Alignment(horizontal="center")
wsSem.cell(row=linha_total, column=3).alignment = Alignment(horizontal="center")

# FORMATAÇÃO DO CABEÇALHO
for row in wsSem.iter_rows(max_row=1, min_col=wsSem.min_column, max_col=wsSem.max_column):
    font_cabecalho = Font(color='00FFFFFF', bold=True, size=14)
    fill_cabecalho = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    for cell in row:
        cell.fill = fill_cabecalho
        cell.font = font_cabecalho

# FORMATAÇÃO DOS DADOS
fill_par = PatternFill(start_color='DCE6F1', end_color='DCE6F1', fill_type='solid')  # azul claro
fill_impar = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')  # branco


for row in wsSem.iter_rows(min_row=2, max_row=wsSem.max_row, min_col=wsSem.min_column, max_col=wsSem.max_column):
    numero_linha = row[0].row  # pega o número da linha
    for cell in row:
        if numero_linha % 2 == 0:
            cell.fill = fill_par
        else:
            cell.fill = fill_impar

    # Bordas
for row in wsSem.iter_rows(max_col=wsSem.max_column, min_row=1, max_row=wsSem.max_row):
    thin = Side(border_style='thin', color='000000')
    for cell in row:
        cell.border = Border(top=thin, bottom=thin, left=thin, right=thin, vertical=thin, horizontal=thin)

    # R$
for row in wsSem.iter_rows(min_col=3, max_col=3, min_row=2, max_row=wsSem.max_row):
    for cell in row:
        cell.number_format = ("R$ " + FORMAT_NUMBER_COMMA_SEPARATED2)
        cell.alignment = Alignment(horizontal="center")

    # __/__/__
for row in wsSem.iter_rows(min_col=4, max_col=4, min_row=2, max_row=wsSem.max_row):
    for cell in row:
        cell.number_format = FORMAT_DATE_DDMMYY
        cell.alignment = Alignment(horizontal="center")

    # Tamanhos das Colunas

wsSem.column_dimensions['A'].width = 30
wsSem.column_dimensions['B'].width = 50
wsSem.column_dimensions['C'].width = 15
wsSem.column_dimensions['D'].width = 15
wsSem.column_dimensions['E'].width = 30
wsSem.column_dimensions['F'].width = 25
wsSem.column_dimensions['G'].width = 35

# Filtro apenas para a primeira tabela (sem incluir a linha TOTAL)
ultima_coluna = get_column_letter(wsSem.max_column)
penultima_linha = linha_total - 1  # isso garante que a linha TOTAL fique fora do filtro
wsSem.auto_filter.ref = f"A1:{ultima_coluna}{penultima_linha}"
  # aplicar filtro somente até a penúltima linha

wbSem.save("RelatoriosProntos\RELATÓRIO_Semanal.xlsx")


