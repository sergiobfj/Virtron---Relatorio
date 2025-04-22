import openpyxl as xl
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.styles.numbers import FORMAT_DATE_DDMMYY, FORMAT_NUMBER_COMMA_SEPARATED2
from openpyxl.styles.alignment import Alignment
from openpyxl.utils import get_column_letter
import pandas as pd

#--------------------------RELATÓRIO SEMANAL----------------------------------------------
#-------------------------------SUNHUB--------------------------------

# Dados SUNHUB
wbSun = xl.load_workbook("ArquivosSunhub\Relatório Diretoria - SUNHUB.xlsx") # < ----------------------
wsSun = wbSun['Ploomes']

# Últimas Linhas e Colunas
maxR = wsSun.max_row
maxC = wsSun.max_column

dados = []
# Percorrendo os dados do SunHub
for row in wsSun.iter_rows(min_row=1, max_row=maxR, max_col=maxC):
    linha = []
    for cell in row:
        linha.append(cell.value)
    dados.append(linha)


#-------------------------------RELATORIO MENSAL--------------------------------


# Criando a Workbook de Relatório
wbRel = xl.Workbook()
wsRel = wbRel.active
wsRel.title = "Mensal"

# Copiando os dados da Lista para a Workbook
for l, linha in enumerate(dados, start=1):  # Começa na linha 1
    for c, valor in enumerate(linha, start=1):  # Começa na coluna 1
        wsRel.cell(row=l, column=c, value=valor)


# LINHA 'TOTAL'
linha_total = wsRel.max_row + 1
wsRel.cell(row=linha_total, column=1).value = "Total:"  # TOTAL
#wsRel.cell(row=linha_total, column=2).value = f'=CONT.VALORES(B2:B{linha_total - 1})'  # SOMA DE VENDAS
#wsRel.cell(row=linha_total, column=3).value = f'=SUM(C2:C{linha_total - 1})'  # SOMA DOS VALORES
wsRel.cell(row=linha_total, column=2).alignment = Alignment(horizontal="center")
wsRel.cell(row=linha_total, column=3).alignment = Alignment(horizontal="center")

# FORMATAÇÃO DO CABEÇALHO
for row in wsRel.iter_rows(max_row=1, min_col=wsRel.min_column, max_col=wsRel.max_column):
    font_cabecalho = Font(color='00FFFFFF', bold=True, size=14)
    fill_cabecalho = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    for cell in row:
        cell.fill = fill_cabecalho
        cell.font = font_cabecalho

# FORMATAÇÃO DOS DADOS
fill_par = PatternFill(start_color='DCE6F1', end_color='DCE6F1', fill_type='solid')  # azul claro
fill_impar = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')  # branco


for row in wsRel.iter_rows(min_row=2, max_row=wsRel.max_row, min_col=wsRel.min_column, max_col=wsRel.max_column):
    numero_linha = row[0].row  # pega o número da linha
    for cell in row:
        if numero_linha % 2 == 0:
            cell.fill = fill_par
        else:
            cell.fill = fill_impar

    # Bordas
for row in wsRel.iter_rows(max_col=wsRel.max_column, min_row=1, max_row=wsRel.max_row):
    thin = Side(border_style='thin', color='000000')
    for cell in row:
        cell.border = Border(top=thin, bottom=thin, left=thin, right=thin, vertical=thin, horizontal=thin)

    # R$
for row in wsRel.iter_rows(min_col=3, max_col=3, min_row=2, max_row=wsRel.max_row):
    for cell in row:
        cell.number_format = ("R$ " + FORMAT_NUMBER_COMMA_SEPARATED2)
        cell.alignment = Alignment(horizontal="center")

    # __/__/__
for row in wsRel.iter_rows(min_col=4, max_col=4, min_row=2, max_row=wsRel.max_row):
    for cell in row:
        cell.number_format = FORMAT_DATE_DDMMYY
        cell.alignment = Alignment(horizontal="center")

    # Tamanhos das Colunas

wsRel.column_dimensions['A'].width = 30
wsRel.column_dimensions['B'].width = 50
wsRel.column_dimensions['C'].width = 15
wsRel.column_dimensions['D'].width = 15
wsRel.column_dimensions['E'].width = 30
wsRel.column_dimensions['F'].width = 25
wsRel.column_dimensions['G'].width = 35

# Filtro apenas para a primeira tabela (sem incluir a linha TOTAL)
ultima_coluna = get_column_letter(wsRel.max_column)
penultima_linha = linha_total - 1  # isso garante que a linha TOTAL fique fora do filtro
wsRel.auto_filter.ref = f"A1:{ultima_coluna}{penultima_linha}"  # aplicar filtro somente até a penúltima linha


#-------------------------------GERENTES--------------------------------

# Carregando os dados dos gerentes
dados_gerentes1 = pd.read_excel("ArquivosSunhub\Planilha_Jose.xlsx")# < ----------------------
dados_gerentes2 = pd.read_excel("ArquivosSunhub\Planilha_Sr.Edilson.xlsx") # < ----------------------
dados_gerentes3 = pd.read_excel("ArquivosSunhub\Planilha_Valter.xlsx") # < ----------------------


# -------------------- CRIAR TABELAS INDIVIDUAIS-----------------------


# Adicionando a tabela dos gerentes
def adicionar_tabela_gerentes(wsRel, dados_gerentes, linha_inicial):
    # Adicionando os dados dos gerentes
    for l, linha in enumerate(dados_gerentes.values, start=linha_inicial + 1):  # dados abaixo do cabeçalho
        for c, valor in enumerate(linha, start=1):
            wsRel.cell(row=l, column=c, value=valor)

    # Cabeçalho dos gerentes
    for c, nome_coluna in enumerate(dados_gerentes.columns, start=1):
        wsRel.cell(row=linha_inicial, column=c, value=nome_coluna)
        cell = wsRel.cell(row=linha_inicial, column=c)
        cell.fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
        cell.font = Font(color='00FFFFFF', bold=True, size=12)

    # Formatação das linhas
    fill_par = PatternFill(start_color='DCE6F1', end_color='DCE6F1', fill_type='solid')
    fill_impar = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    
    linha_final = linha_inicial + len(dados_gerentes)

    for row in wsRel.iter_rows(min_row=linha_inicial + 1, max_row=linha_final, min_col=1, max_col=dados_gerentes.shape[1]):
        numero_linha = row[0].row
        for cell in row:
            cell.fill = fill_par if numero_linha % 2 == 0 else fill_impar

    # Fórmulas de Total para os gerentes
    linha_total_gerentes = linha_final + 1
    wsRel.cell(row=linha_total_gerentes, column=1, value="Total:")

    #wsRel.cell(row=linha_total_gerentes, column=2).value = f'=CONT.VALORES(B{linha_inicial + 1}:B{linha_final})'
    #wsRel.cell(row=linha_total_gerentes, column=3).value = f'=SOMA(C{linha_inicial + 1}:C{linha_final})'
    wsRel.cell(row=linha_total_gerentes, column=2).alignment = Alignment(horizontal="center")
    wsRel.cell(row=linha_total_gerentes, column=3).alignment = Alignment(horizontal="center")

    # Formatação dos valores (R$ e datas)
    for row in wsRel.iter_rows(min_row=linha_inicial + 1, max_row=linha_total_gerentes, min_col=3, max_col=3):
        for cell in row:
            cell.number_format = ("R$ " + FORMAT_NUMBER_COMMA_SEPARATED2)
            cell.alignment = Alignment(horizontal="center")

    for row in wsRel.iter_rows(min_row=linha_inicial + 1, max_row=linha_total_gerentes, min_col=4, max_col=4):
        for cell in row:
            cell.number_format = FORMAT_DATE_DDMMYY
            cell.alignment = Alignment(horizontal="center")

    # Bordas para toda a área dos gerentes
    for row in wsRel.iter_rows(min_row=linha_inicial, max_row=linha_total_gerentes, min_col=1, max_col=dados_gerentes.shape[1]):
        for cell in row:
            thin = Side(border_style='thin', color='000000')
            cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)



# -------------------- CRIAR ABAS INDIVIDUAIS -----------------------


def criar_aba_individual(wb, nome_aba, dados_gerente):
    ws = wb[nome_aba]
    
    # Adicionando cabeçalho
    for c, nome_coluna in enumerate(dados_gerente.columns, start=1):
        cell = ws.cell(row=1, column=c, value=nome_coluna)
        cell.fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
        cell.font = Font(color='00FFFFFF', bold=True, size=12)
    
    # Adicionando os dados
    for l, linha in enumerate(dados_gerente.values, start=2):
        for c, valor in enumerate(linha, start=1):
            ws.cell(row=l, column=c, value=valor)

    # Formatação de linhas alternadas
    fill_par = PatternFill(start_color='DCE6F1', end_color='DCE6F1', fill_type='solid')
    fill_impar = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=dados_gerente.shape[1]):
        for cell in row:
            cell.fill = fill_par if cell.row % 2 == 0 else fill_impar

    # Formatando valores e datas
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=3, max_col=3):
        for cell in row:
            cell.number_format = ("R$ " + FORMAT_NUMBER_COMMA_SEPARATED2)
            cell.alignment = Alignment(horizontal="center")
    
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=4, max_col=4):
        for cell in row:
            cell.number_format = FORMAT_DATE_DDMMYY
            cell.alignment = Alignment(horizontal="center")

    # Bordas
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=dados_gerente.shape[1]):
        for cell in row:
            thin = Side(border_style='thin', color='000000')
            cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)
    
    # Auto largura

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 25
    ws.column_dimensions['G'].width = 35


# -------------------- ADICIONAR DADOS NO MENSAL + CRIAR ABAS INDIVIDUAIS -----------------------

linha_inicial_gerentes = wsRel.max_row + 2
adicionar_tabela_gerentes(wsRel, dados_gerentes1, linha_inicial_gerentes)
ws1 = wbRel.create_sheet('Jose')
criar_aba_individual(wbRel, 'Jose', dados_gerentes1)


linha_inicial_gerentes = wsRel.max_row + 2
adicionar_tabela_gerentes(wsRel, dados_gerentes2, linha_inicial_gerentes)
ws2 = wbRel.create_sheet('Sr.Edilson')
criar_aba_individual(wbRel, 'Sr.Edilson', dados_gerentes2)


linha_inicial_gerentes = wsRel.max_row + 2
adicionar_tabela_gerentes(wsRel, dados_gerentes3, linha_inicial_gerentes)
ws3 = wbRel.create_sheet('Valter')
criar_aba_individual(wbRel, 'Valter', dados_gerentes3)

abas_gerentes = [ws1, ws2, ws3]

for aba in abas_gerentes:
    linha_inicial = 1
    linha_final = aba.max_row
    linha_total_gerentes = linha_final + 1

    aba.cell(row=linha_total_gerentes, column=1).value = "Total:"  # TOTAL
    aba.cell(row=linha_total_gerentes, column=2).value = f'=CONT.VALORES(B{linha_inicial + 1}:B{linha_final})'
    aba.cell(row=linha_total_gerentes, column=3).value = f'=SOMA(C{linha_inicial + 1}:C{linha_final})'
    aba.cell(row=linha_total_gerentes, column=2).alignment = Alignment(horizontal="center")
    aba.cell(row=linha_total_gerentes, column=3).alignment = Alignment(horizontal="center")


# Salvando a planilha final
wbRel.save("RelatoriosProntos\RELATÓRIO_Mensal.xlsx")
